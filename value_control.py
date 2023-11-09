import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

# Podfunkce pro zpracování dat z Excelu
def read_excel(file_path: str) -> list:
    test_data: list = []
    df = pd.read_excel(file_path, sheet_name="List 1", skiprows=0)
    # Iterate through the rows of the DataFrame and print the content of each cell
    for index, row in df.iterrows():
        line: list = []
        for column in df.columns:
            cell_content = str(row[column])
            line.append(cell_content)
        test_data.append(line)

    return test_data

# Podfukce pro zpracování dat z HTML souboru
def read_html(file_path: str) -> list:
    # Set pandas options to display more rows and the full content of columns
    pd.set_option('display.max_rows', None)  # Set to None to display all rows, or a large number
    pd.set_option('display.max_colwidth', None)  # Set to None to display full column width
    pd.set_option('display.width', None)

    df = pd.read_html(file_path)

    # Concatenate contents of all tables into a single DataFrame
    concatenated_df = pd.concat(df, ignore_index=True)
    list_df: list = concatenated_df.values.tolist()
    list_df.pop(0)  # Removal of description

    # Filtration of needed info [Energy, Pressure, Amplitude]
    for row in list_df:
        for integer, value in enumerate(row):
            if integer in range(9):
                row.pop(1)
        for integer, value in enumerate(row):
            if integer in range(5):
                row.pop(4)
        row.pop(4)
        # value swap so that it corresponds with Excel data
        element: str = row.pop(1)
        row.append(element)

    return list_df

# Podfukce pro procentuální výpočet
def percentage_calculation(html_value: str, excel_value: str) -> str:
    if float(excel_value) > (float(html_value)*1.2):
        return "ABOVE"
    elif float(excel_value) < (float(html_value)*0.8):
        return "BELOW"
    else:
        return "OKAY"

# Podfunkce pro vytvoření nového Excelového souboru se zaznamenanými výsledky
def save_new_excel(file_path: str, original_excel) -> None:
    new_file = file_path
    original_excel.save(new_file)

# Podfunkce pro zápis výsledků porovnávání do Excelu
def excel_test_record(results: str, row: str, work_sheet) -> None:

    cell_red: str = "ff0000"
    cell_green: str = "3cb371"
    cell_orange: str = "ffa500"

    for x, y in enumerate(["G", "H", "I"], 1):
        if results[x] == "ABOVE":
            work_sheet[f"A{y}{int(float(row)) + 1}"] = "ABOVE"
            color = PatternFill(start_color=cell_red, end_color=cell_red, fill_type="solid")
        elif results[x] == "BELOW":
            work_sheet[f"A{y}{int(float(row)) + 1}"] = "BELOW"
            color = PatternFill(start_color=cell_orange, end_color=cell_orange, fill_type="solid")
        elif results[x] == "OKAY":
            work_sheet[f"A{y}{int(float(row)) + 1}"] = "OKAY"
            color = PatternFill(start_color=cell_green, end_color=cell_green, fill_type="solid")
        else:
            work_sheet[f"A{y}{int(float(row)) + 1}"] = "NO DATA"
            color = PatternFill(start_color="7c0000", end_color="7c0000", fill_type="solid")

        cell = work_sheet[f"A{y}{int(float(row)) + 1}"]
        cell.fill = color

# Podfunkce pro porovnávání testovacích parametrů předepsaných v HTML vs skutečných v Excelu
def data_comparison(file_path: str, html_data: list, test_data: list) -> None:
    original_excel = openpyxl.load_workbook(file_path)
    work_sheet = original_excel['List 1']
    for html_row in html_data:
        print(html_row)
        for excel_row in test_data:
            result_verdikt: list = []
            if html_row[0] == excel_row[3]:
                print(excel_row)
                pressure_result: bool = percentage_calculation(html_row[1], excel_row[10])
                amp_result: bool = percentage_calculation(html_row[2], excel_row[11])
                energy_result: bool = percentage_calculation(html_row[3], excel_row[12])
                result_verdikt.append(html_row[0])
                result_verdikt.append(pressure_result)
                result_verdikt.append(amp_result)
                result_verdikt.append(energy_result)
                excel_test_record(results=result_verdikt, row=excel_row[0], work_sheet=work_sheet)
                print(f"Results: {result_verdikt}")
        print()
    save_new_excel(file_path="Checked_file.xlsx", original_excel=original_excel)

# Hlavní fuknce pro management podfunkci
def main() -> None:
    html_data: list = read_html(file_path="SAP.MHTML")
    test_data: list = read_excel(file_path="Original_file.xlsx")
    data_comparison(file_path="Original_file.xlsx", html_data=html_data, test_data=test_data)

# Spouštěč Programu
if __name__ == '__main__':
    main()
